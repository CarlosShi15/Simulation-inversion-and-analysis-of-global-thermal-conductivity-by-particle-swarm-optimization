import os

import numpy as np
import matplotlib.pyplot as plt
from tqdm import tqdm
from multiprocessing import Pool
from openpyxl import Workbook

# Function to evolve the heat distribution
def evolve(u, u_previous, dt, dx2, dy2, alpha):
    nx, ny = u.shape
    for i in range(1, nx - 1):
        for j in range(1, ny - 1):
            u[i, j] = u_previous[i, j] + alpha * dt * (
                (u_previous[i + 1, j] - 2 * u_previous[i, j] + u_previous[i - 1, j]) / dx2 +
                (u_previous[i, j + 1] - 2 * u_previous[i, j] + u_previous[i, j - 1]) / dy2)
    return u

# Function to initialize fields
def init_fields(lenX, lenY, Tguess, Ttop, Tbottom, Tleft, Tright):
    field = np.empty((lenX, lenY), dtype=np.float64)
    field.fill(Tguess)
    field[(lenY - 1):, :] = Ttop
    field[:1, :] = Tbottom
    field[:, (lenX - 1):] = Tright
    field[:, :1] = Tleft
    field0 = field.copy()  # Previous temperature field
    return field, field0

# Function to write field to an image
def write_field(X, Y, field, step, size, device, lenX, timesteps):
    plt.gca().clear()
    plt.title("Temperature")
    plt.contourf(X, Y, field, levels=50, cmap=plt.cm.jet)
    if step == 0:
        plt.colorbar()
    plt.savefig(f'heat_{size}_{device}_{lenX}_{timesteps}_{step}.png')

# Function to write field data to an Excel file
def write_to_excel(lenX, lenY, field, step, timesteps, filename):
    workbook = Workbook()
    worksheet = workbook.active

    # Write headers
    worksheet.cell(row=1, column=1, value="Step")
    for i in range(1, lenX + 1):
        worksheet.cell(row=1, column=i + 1, value=f"X={i-1}")

    # Write temperature data
    for j in range(1, lenY + 1):
        worksheet.cell(row=j + 1, column=1, value=j - 1)
        for i in range(1, lenX + 1):
            worksheet.cell(row=j + 1, column=i + 1, value=field[i - 1, j - 1])
    workbook.save(filename)

def simulate(args):
    lenX, lenY, timesteps, image_interval, size, device = args

    # Basic parameters
    k = 180  # Thermal conductivity(6061 o态铝合金)
    c = 900  # Specific heat capacity
    rho = 2710  # Density
    alpha = k / (c * rho)  # Thermal diffusivity

    # Grid spacing
    dx = 0.01
    dy = 0.01
    dx2 = dx ** 2
    dy2 = dy ** 2

    # Time step
    dt = 0.01  # Example time step, adjust as necessary

    # Set up grid
    X, Y = np.meshgrid(np.arange(0, lenX), np.arange(0, lenY))

    # Boundary conditions
    Ttop = 20
    Tbottom = 20
    Tleft = 20
    Tright = 20

    # Initial guess for internal temperature field
    Tguess = 100

    # Initialize fields
    field, field0 = init_fields(lenX, lenY, Tguess, Ttop, Tbottom, Tleft, Tright)

    # Simulation loop
    for m in tqdm(range(1, timesteps + 1)):
        field = evolve(field, field0, dt, dx2, dy2, alpha)
        field0 = field.copy()

        if m % image_interval == 0:
            write_field(X, Y, field, m, size, device, lenX, timesteps)
            write_to_excel(lenX, lenY, field, m, timesteps, f'temperature_{size}_{device}_{lenX}_{timesteps}_{m}.xlsx')


    # Write final temperature field
    write_field(X, Y, field, timesteps, size, device, lenX, timesteps)
    write_to_excel(lenX, lenY, field, timesteps, timesteps, f'temperature_{size}_{device}_{lenX}_{timesteps}_final.xlsx')




if __name__ == '__main__':
    lenX = 20
    lenY = 20
    timesteps = 100
    image_interval = 1000
    size = 1
    device = 'cpu'

    # Parameters for parallel execution
    num_processes = 4 # Adjust according to your system
    pool = Pool(processes=num_processes)

    # Split the total number of timesteps evenly among processes
    timesteps_per_process = timesteps // num_processes

    # Create argument list for parallel execution
    args_list = [(lenX, lenY, timesteps_per_process, image_interval, size, device)] * num_processes

    # Run simulations in parallel
    pool.map(simulate, args_list)

    # Close the pool of processes
    pool.close()
    pool.join()
