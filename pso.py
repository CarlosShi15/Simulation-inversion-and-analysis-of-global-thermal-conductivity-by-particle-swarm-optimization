import os
import numpy as np
from tqdm import tqdm
from openpyxl import load_workbook
from scipy.optimize import minimize
import numpy as np
import matplotlib.pyplot as plt
from tqdm import tqdm
from multiprocessing import Pool
from openpyxl import Workbook
import numpy as np
import random

# Function to evolve the heat distribution
def evolve(u, u_previous, dt, dx2, dy2, alpha):
    nx, ny = u.shape
    for i in range(1, nx - 1):
        for j in range(1, ny - 1):
            u[i, j] = u_previous[i, j] + alpha * dt * (
                    (u_previous[i + 1, j] - 2 * u_previous[i, j] + u_previous[i - 1, j]) / dx2 +
                    (u_previous[i, j + 1] - 2 * u_previous[i, j] + u_previous[i, j - 1]) / dy2)
    return u


# Function to initialize fields
def init_fields(lenX, lenY, Tguess, Ttop, Tbottom, Tleft, Tright):
    field = np.empty((lenX, lenY), dtype=np.float64)
    field.fill(Tguess)
    field[(lenY - 1):, :] = Ttop
    field[:1, :] = Tbottom
    field[:, (lenX - 1):] = Tright
    field[:, :1] = Tleft
    field0 = field.copy()  # Previous temperature field
    return field, field0


# Function to read real temperature data from an Excel file
def read_real_data(filename):
    workbook = load_workbook(filename)
    sheet = workbook.active
    data = []
    for row in sheet.iter_rows(min_row=2, values_only=True):  # Start from the second row
        row_data = list(row)[1:]  # Skip the first column
        data.append(row_data)

    # Convert the list to a 15x15 NumPy array
    data_array = np.array(data)
    fixed_array = np.zeros((20, 20))
    fixed_array[:data_array.shape[0], :data_array.shape[1]] = data_array

    return fixed_array


# Function to calculate the error between simulated and real temperature data
def calculate_error(simulated, real):
    return np.sqrt(np.mean((simulated - real) ** 2))


def simulate(args):
    lenX, lenY, timesteps, image_interval, size,k = args

    # Basic parameters
    c = 900  # Specific heat capacity
    rho = 2710  # Density
    alpha = k / (c * rho)  # Thermal diffusivity

    # Grid spacing
    dx = 0.8
    dy = 0.8
    dx2 = dx ** 2
    dy2 = dy ** 2

    # Time step
    dt = 0.8  # Example time step, adjust as necessary

    # Set up grid
    X, Y = np.meshgrid(np.arange(0, lenX), np.arange(0, lenY))

    # Boundary conditions
    Ttop = 20
    Tbottom = 20
    Tleft = 20
    Tright = 20

    # Initial guess for internal temperature field
    Tguess = 100

    # Initialize fields
    field, field0 = init_fields(lenX, lenY, Tguess, Ttop, Tbottom, Tleft, Tright)

    # Simulation loop
    for m in tqdm(range(1, timesteps + 1)):
        field = evolve(field, field0, dt, dx2, dy2, alpha)
        field0 = field.copy()

    # Return final temperature field
    return field


def PSO(num_particles, num_iterations, lenX, lenY, timesteps_per_process, image_interval, size, initial_k, real_data):
    # Initialize particle positions and velocities
    positions = [initial_k + random.uniform(-1, 1) for _ in range(num_particles)]
    velocities = [random.uniform(-1, 1) for _ in range(num_particles)]

    # Initialize global best position and error
    global_best_position = positions[0]
    global_best_error = float('inf')

    for _ in range(num_iterations):
        for i in range(num_particles):
            # Update velocity
            inertia = 0.5
            cognitive_weight = 1
            social_weight = 2
            r1 = random.random()
            r2 = random.random()
            velocities[i] = inertia * velocities[i] + cognitive_weight * r1 * (
                        positions[i] - global_best_position) + social_weight * r2 * (
                                        global_best_position - positions[i])

            # Update position
            positions[i] += velocities[i]
            positions[i] = round(positions[i])  # Ensure k is an integer

            # Simulate and calculate error
            simulated_data = simulate((lenX, lenY, timesteps_per_process, image_interval, size, positions[i]))
            error = calculate_error(simulated_data, real_data)

            # Update global best
            if error < global_best_error:
                global_best_error = error
                global_best_position = positions[i]

    return global_best_position



if __name__ == '__main__':
    num_particles = 10
    num_iterations = 20
    lenX = 100
    lenY = 100
    timesteps_per_process = 10000
    image_interval = 1000
    size = 1
    initial_k = 100
    real_data = read_real_data('temperature_1_cpu_20_25_final.xlsx')

    best_k = PSO(num_particles, num_iterations, lenX, lenY, timesteps_per_process, image_interval, size, initial_k,
                 real_data)
    print(f"Optimal k value: {best_k}")


